from selenium import webdriver
from bs4 import BeautifulSoup
from selenium.webdriver.common.by import By
import time
from openpyxl import load_workbook

driver = webdriver.Chrome('./chromedriver.exe')
#크롬 드라이버에 url 주소 넣고 실행
workbook=load_workbook("C:/Users/82109/Desktop/데이터 수집/학교공지사항.xlsx")

# 페이지가 완전히 로딩되도록 3초동안 기다림
driver.get('http://board.sejong.ac.kr/boardlist.do?bbsConfigFK=333')

for i in range(2,11):
    time.sleep(3)
    html = driver.page_source # 페이지의 elements모두 가져오기
    soup = BeautifulSoup(html, 'html.parser')
    ul = soup.select_one('table.text-board')
    elem=ul.select('body > div > table > tbody> tr')
    
    datas=[]
    for tr in elem:
        title=tr.select_one('td.subject').get_text().replace('\n', '').replace('\t', '')
        url="https://"+tr.select_one('td.subject > a').get('href').replace('/viewcount.do?rtnUrl=','board.sejong.ac.kr').replace('^','&')
        date=tr.select_one('td.date').get_text()
        datas.append([title,url,date])
    
    worksheet = workbook['Sheet1']
    for data in datas:
        worksheet.append(data)

    workbook.save("C:/Users/82109/Desktop/데이터 수집/학교공지사항.xlsx")
   
    nextpage = driver.find_element(By.CSS_SELECTOR,f"body > div > div.pagination > ul > span.num > a:nth-child({i})")
    nextpage.click()
    
    
# import requests
# from bs4 import BeautifulSoup
# from openpyxl import load_workbook


# url = 'http://board.sejong.ac.kr/boardlist.do?bbsConfigFK=333'

# response = requests.get(url)
# workbook=load_workbook("C:/Users/82109/Desktop/데이터 수집/학교공지사항.xlsx")

# if response.status_code == 200:
#     html = response.text
#     ##제목 가져오기
    # soup = BeautifulSoup(html, 'html.parser')
    # ul = soup.select_one('table.text-board')
    # elem=ul.select('tbody>tr')
    
    # datas=[]
    # for tr in elem:
    #     title=tr.select_one('td.subject').get_text().replace('\n', '').replace('\t', '')
    #     url=tr.select_one('td.subject > a').get('href').replace('/viewcount.do?rtnUrl=','board.sejong.ac.kr').replace('^','&')
    #     date=tr.select_one('td.date').get_text()
    #     datas.append([title,url,date])
    
    
    # worksheet = workbook['결과']
    # for data in datas:
    #     worksheet.append(data)

    # workbook.save("C:/Users/82109/Desktop/데이터 수집/학교공지사항.xlsx")




# else : 
#     print(response.status_code)





# from openpyxl import Workbook

#         # 엑셀파일 쓰기
# write_wb = Workbook()

#         # 이름이 있는 시트를 생성
# write_ws = write_wb.create_sheet('생성시트')

#         # Sheet1에다 입력
# write_ws = write_wb.active
# write_ws['A1'] = '숫자'

#         #행 단위로 추가
# write_ws.append([1,2,3])

#         #셀 단위로 추가
# write_ws.cell(5, 5, '5행5열')

# ##엑셀파일 저장
# write_wb.save("C:/Users/82109/Desktop/데이터 수집/숫자.xlsx")